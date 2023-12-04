def master():
    from openpyxl import load_workbook
    wi = load_workbook(filename= 'Route Master.xlsx')
    sheet_ranges = wi['route']
    maxi = sheet_ranges.max_row
    setM = {}
    f = open("c.txt","a")
    for i in range(3 , maxi+1):
        print(setM.add(sheet_ranges[f"E{i}"].value))

master()